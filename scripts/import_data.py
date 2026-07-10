"""
JIL CRM Data Import Script
Run:  python odoo-bin shell -c odoo.conf -d odoo_jil_crom < scripts/import_data.py
"""
import openpyxl
import logging

_logger = logging.getLogger(__name__)

XLSX_PATH = r'C:\Users\OTHMANE\Desktop\Projet\Odoo CRM AI - Copie\Data.xlsx'

PARTNER_SHEETS = {'Clients': 'customer', 'Fournisseurs': 'supplier'}

PRODUCT_CATEGORY_SHEETS = [
    'Articles ADS', 'Articles Bouchon Couronne', 'Articles Consignation',
    'Articles Frais Gestion', 'Articles Metal', 'Articles Plastique',
    'Articles Pres Serv', 'Articles Twist Off', 'Article Verre',
]


def _get_ws(sheet_name):
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True, read_only=True)
    return wb[sheet_name] if sheet_name in wb.sheetnames else None


def import_partners(env):
    for sheet_name, cat_type in PARTNER_SHEETS.items():
        ws = _get_ws(sheet_name)
        if not ws:
            continue
        count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            ref, name = (row + (None, None))[:2]
            if ref and name:
                ref_s = str(ref).strip()
                name_s = str(name).strip()
                if not env['res.partner'].search([('ref', '=', ref_s)], limit=1):
                    env['res.partner'].create({
                        'ref': ref_s,
                        'name': name_s,
                        'company_type': 'company',
                        'customer_rank': 1 if cat_type == 'customer' else 0,
                        'supplier_rank': 1 if cat_type == 'supplier' else 0,
                    })
                    count += 1
        _logger.info("Imported %s %s", count, sheet_name)


def import_products(env):
    for sheet_name in PRODUCT_CATEGORY_SHEETS:
        ws = _get_ws(sheet_name)
        if not ws:
            continue
        fam = sheet_name.replace('Articles ', '').replace('Article ', '')
        category = env['product.category'].search([('name', '=', fam)], limit=1)
        if not category:
            category = env['product.category'].create({'name': fam})
        count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            row = (row + (None,) * 5)[:5]
            ptype, ref, name = row[0], row[1], row[2]
            if ref and name:
                ref_s = str(ref).strip()
                name_s = str(name).strip()
                if not env['product.product'].search([('default_code', '=', ref_s)], limit=1):
                    env['product.product'].create({
                        'default_code': ref_s,
                        'name': name_s,
                        'type': 'consu',
                        'categ_id': category.id,
                    })
                    count += 1
        _logger.info("Imported %s products in %s", count, sheet_name)


def import_chart_of_accounts(env):
    ws = _get_ws('Plan Comptable')
    if not ws:
        return
    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        row = (row + (None,) * 3)[:3]
        atype, code, name = row
        if code and name:
            code_s, name_s = str(code).strip(), str(name).strip()
            if not env['account.account'].search([('code', '=', code_s)], limit=1):
                env['account.account'].create({
                    'code': code_s,
                    'name': name_s,
                    'account_type': 'asset_receivable' if atype and 'client' in str(atype).lower() else 'asset_current',
                    'reconcile': True,
                })
                count += 1
    _logger.info("Imported %s chart of accounts", count)


def run(env):
    _logger.info("Starting JIL CRM data import from %s", XLSX_PATH)
    import_partners(env)
    import_products(env)
    import_chart_of_accounts(env)
    env.cr.commit()
    _logger.info("Data import completed")

run(env)
